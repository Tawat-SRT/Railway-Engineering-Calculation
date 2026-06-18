from __future__ import annotations

from datetime import date
from io import BytesIO
from math import ceil

import pandas as pd
import streamlit as st
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill

from calculators import (
    RAIL_PROPERTIES,
    cant_calculation,
    drainage_design,
    improve_string_lining_versines,
    sleeper_track_analysis,
    string_lining_calculation,
    theoretical_versine_profile,
    track_widening,
    transition_curve,
    turnout_geometry,
    turnout_speed,
)


st.set_page_config(
    page_title="SRT Railway Calculator",
    page_icon="🛤️",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown(
    """
<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+Thai:wght@400;500;600;700&family=Roboto+Mono:wght@500&display=swap');
:root { --navy:#0B2341; --maroon:#7A0019; --cream:#F7F3EA; --ink:#182333; --muted:#667085; }
html, body, [class*="css"] { font-family:'Noto Sans Thai',sans-serif; }
[data-testid="stAppViewContainer"] { background:#fbfaf7; color:var(--ink); }
[data-testid="stSidebar"] { background:var(--navy); border-right:1px solid rgba(255,255,255,.08); }
[data-testid="stSidebar"] * { color:#f8fafc !important; }
[data-testid="stSidebar"] [role="radiogroup"] label { padding:.52rem .65rem; border-radius:7px; }
[data-testid="stSidebar"] [role="radiogroup"] label:has(input:checked) { background:rgba(255,255,255,.12); }
.block-container { padding-top:2.2rem; max-width:1500px; }
h1,h2,h3 { color:var(--navy); letter-spacing:-.02em; }
h1 { font-size:2.15rem !important; margin-bottom:.15rem !important; }
.eyebrow { color:var(--maroon); font-weight:700; letter-spacing:.08em; text-transform:uppercase; font-size:.78rem; }
.subhead { color:var(--muted); font-size:1rem; margin-bottom:1.45rem; }
.rail-rule { height:8px; margin:0 0 1.3rem; background:repeating-linear-gradient(90deg,var(--maroon) 0 54px,transparent 54px 66px); border-top:2px solid var(--navy); }
.section-label { font-size:.77rem; color:var(--muted); font-weight:700; letter-spacing:.06em; text-transform:uppercase; margin:1rem 0 .2rem; }
.result-band { border-top:1px solid #d8d5cf; border-bottom:1px solid #d8d5cf; padding:.5rem 0 .25rem; margin:.4rem 0 1rem; }
.ok { color:#146c43; font-weight:700; }
.bad { color:#a61b1b; font-weight:700; }
.formula { font-family:'Roboto Mono',monospace; color:#344054; background:#f1eee8; padding:.7rem .85rem; border-left:3px solid var(--maroon); font-size:.84rem; }
.catalog-row { padding:.8rem 0; border-bottom:1px solid #e4e1db; }
.catalog-row b { color:var(--navy); }
div[data-testid="stMetric"] { border-left:3px solid var(--maroon); padding-left:.85rem; }
div[data-testid="stMetric"] label { color:#667085; }
.stButton>button, .stDownloadButton>button { border-radius:6px; border:1px solid var(--navy); font-weight:650; }
.stDownloadButton>button { background:var(--navy); color:white; }
[data-testid="stNumberInput"] input, [data-testid="stSelectbox"] > div > div { background:white; }
@media(max-width:768px){ .block-container{padding:1.2rem 1rem;} h1{font-size:1.7rem!important;} }
</style>
""",
    unsafe_allow_html=True,
)


PAGES = {
    "ภาพรวมเครื่องมือ": "overview",
    "ค่ายกโค้งและความเร็ว": "cant",
    "ความเร็วผ่านประแจ": "turnout_speed",
    "ออกแบบทางระบายน้ำ": "drainage",
    "วิเคราะห์ราง–หมอน–หินโรยทาง": "sleeper",
    "เรขาคณิตประแจ": "turnout_geometry",
    "โค้งต่อและโค้งกลม": "transition",
    "โค้งเส้นเชือก (String Lining)": "string_lining",
    "การขยายขนาดทาง": "widening",
}


def heading(title: str, subtitle: str, eyebrow: str = "Railway Engineering Calculator") -> None:
    st.markdown(f'<div class="eyebrow">{eyebrow}</div>', unsafe_allow_html=True)
    st.title(title)
    st.markdown(f'<div class="subhead">{subtitle}</div><div class="rail-rule"></div>', unsafe_allow_html=True)


def result_status(label: str, ok: bool, detail: str) -> None:
    cls, text = ("ok", "ผ่านเกณฑ์") if ok else ("bad", "ต้องตรวจสอบ")
    st.markdown(f"**{label}** — <span class='{cls}'>{text}</span> · {detail}", unsafe_allow_html=True)


def string_lining_curve_diagram(direction: str = "โค้งขวา") -> None:
    mirror = "translate(900 0) scale(-1 1)" if direction == "โค้งซ้าย" else ""
    st.markdown(
        f"""
<svg viewBox="0 0 900 260" role="img" aria-label="แผนภาพโค้งทางรถไฟแสดงจุด TS SC CS ST" style="width:100%;height:auto;background:#f5f2eb;border:1px solid #ded9cf;border-radius:8px">
  <defs><filter id="shadow"><feDropShadow dx="0" dy="2" stdDeviation="2" flood-opacity=".18"/></filter></defs>
  <g transform="{mirror}">
    <path d="M45 202 L150 202 C245 202 278 167 335 125 C415 66 505 66 585 125 C642 167 675 202 770 202 L855 202" fill="none" stroke="#0B2341" stroke-width="12" opacity=".16"/>
    <path d="M45 194 L150 194 C245 194 278 159 335 117 C415 58 505 58 585 117 C642 159 675 194 770 194 L855 194" fill="none" stroke="#555" stroke-width="3"/>
    <path d="M45 210 L150 210 C245 210 278 175 335 133 C415 74 505 74 585 133 C642 175 675 210 770 210 L855 210" fill="none" stroke="#555" stroke-width="3"/>
    <path d="M45 202 L150 202 C245 202 278 167 335 125 C415 66 505 66 585 125 C642 167 675 202 770 202 L855 202" fill="none" stroke="#7A0019" stroke-width="2" stroke-dasharray="6 6"/>
    <g fill="#7A0019" stroke="white" stroke-width="3" filter="url(#shadow)">
      <circle cx="150" cy="202" r="8"/><circle cx="335" cy="125" r="8"/><circle cx="585" cy="125" r="8"/><circle cx="770" cy="202" r="8"/>
    </g>
  </g>
  <g font-family="Noto Sans Thai, sans-serif" font-weight="700" fill="#0B2341" text-anchor="middle">
    <text x="150" y="238">TS</text><text x="335" y="101">SC</text><text x="585" y="101">CS</text><text x="770" y="238">ST</text>
  </g>
  <g font-family="Noto Sans Thai, sans-serif" font-size="13" fill="#667085" text-anchor="middle">
    <text x="242" y="232">โค้งต่อด้านต้น</text><text x="460" y="42">โค้งกลม</text><text x="678" y="232">โค้งต่อด้านปลาย</text>
  </g>
</svg>
""",
        unsafe_allow_html=True,
    )


def excel_download(title: str, inputs: dict, results: dict) -> None:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame({"รายการ": list(inputs), "ค่า": list(inputs.values())}).to_excel(writer, "Input", index=False)
        pd.DataFrame({"ผลลัพธ์": list(results), "ค่า": list(results.values())}).to_excel(writer, "Result", index=False)
        ws = writer.book["Result"]
        ws.freeze_panes = "A2"
        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 22
    st.download_button(
        "ดาวน์โหลดผลคำนวณ Excel",
        output.getvalue(),
        file_name=f"{title}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


def string_lining_excel_download(input_df: pd.DataFrame, result_df: pd.DataFrame, baseline: float, metadata: dict) -> None:
    output = BytesIO()
    export = pd.DataFrame(
        {
            "หมุดที่": result_df["หมุดที่"],
            "เวอร์ไซน์เดิม (มม.)": result_df["เวอร์ไซน์เดิม (มม.)"],
            "เวอร์ไซน์ใหม่ (มม.)": result_df["เวอร์ไซน์ใหม่ (มม.)"],
            "ผลต่าง (มม.)": result_df["ผลต่าง (มม.)"],
            "รวมต่าง (มม.)": result_df["รวมต่าง (มม.)"],
            "ครึ่งดัด (มม.)": result_df["ครึ่งดัด (มม.)"],
            "ระยะดัด (มม.)": result_df["ระยะดัด (มม.)"],
            "ราง-ศก. (มม.)": result_df["ราง-ศก. (มม.)"],
            "ค่าขยายทาง (มม.)": input_df["ค่าขยายทาง (มม.)"],
            "ค่ายกโค้ง (มม.)": input_df["ค่ายกโค้ง (มม.)"],
            "เวอร์ไซน์มาตรฐาน (มม.)": input_df["เวอร์ไซน์มาตรฐาน (มม.)"],
            "สทล. จุดที่": input_df["สทล. จุดที่"],
            "ศก.ทาง": input_df["ศก.ทาง"],
            "รางนอก": input_df["รางนอก"],
            "หมายเหตุ": input_df["หมายเหตุ"],
        }
    )
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export.to_excel(writer, "String Lining", index=False, startrow=8)
        ws = writer.book["String Lining"]
        ws.merge_cells("A1:O1")
        ws["A1"] = "รายการคำนวณดัดโค้งด้วยเส้นเชือก (String Lining)"
        ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor="7A0019")
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A2"] = f'โค้งเลขที่ {metadata.get("curve_no", "")}  กม. {metadata.get("km_start", "")} - {metadata.get("km_end", "")}  {metadata.get("direction", "")}'
        ws.merge_cells("A2:O2")
        ws["A3"] = f'ตอนนายตรวจทาง {metadata.get("inspector_section", "")}  แขวง {metadata.get("division", "")}  กอง {metadata.get("office", "")}'
        ws.merge_cells("A3:O3")
        ws["A4"] = f'ความเร็ว {metadata.get("speed", "")} กม./ชม.  รัศมี {metadata.get("radius", "")} ม.  ความยาวโค้ง {metadata.get("curve_length", "")} ม.  ค่ายกโค้ง {metadata.get("cant", "")} มม.'
        ws.merge_cells("A4:O4")
        ws["A5"] = f'ผู้สำรวจ {metadata.get("surveyor", "")}  วันที่ {metadata.get("survey_date", "")}  ผู้คำนวณ {metadata.get("calculator", "")}  วันที่ {metadata.get("calculation_date", "")} '
        ws.merge_cells("A5:O5")
        ws["A6"] = f'ผู้ให้ ศก.หมุด {metadata.get("centre_setter", "")}  ผู้ดัดราง {metadata.get("lining_operator", "")} '
        ws.merge_cells("A6:O6")
        ws.merge_cells("B8:C8")
        ws["B8"] = "เวอร์ไซน์ (มม.)"
        ws.merge_cells("I8:K8")
        ws["I8"] = "ค่ามาตรฐาน (มม.)"
        ws.merge_cells("L8:N8")
        ws["L8"] = "ข้อมูลศูนย์กลางทาง"
        for cell in [ws["B8"], ws["I8"], ws["L8"]]:
            cell.fill = PatternFill("solid", fgColor="D9E2F3")
            cell.font = Font(bold=True, color="0B2341")
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A10"
        ws.auto_filter.ref = f"A9:O{len(export) + 9}"
        ws["Q1"] = "ข้อมูลควบคุม"
        ws["Q2"] = "ระยะฐานราง-ศก. (มม.)"
        ws["R2"] = baseline
        ws["Q4"] = "สูตรจาก Excel ต้นฉบับ"
        ws["Q5"] = "D = B-C"
        ws["Q6"] = "E(i) = E(i-1)+D(i)"
        ws["Q7"] = "F(i) = F(i-1)+E(i-1)"
        ws["Q8"] = "G = 2F"
        ws["Q9"] = "H = Baseline-G"
        first_data_row = 10
        last_data_row = len(export) + 9
        for row in range(first_data_row, last_data_row + 1):
            ws[f"D{row}"] = f"=B{row}-C{row}"
            ws[f"E{row}"] = f"=D{row}" if row == first_data_row else f"=E{row-1}+D{row}"
            ws[f"F{row}"] = "=0" if row == first_data_row else f"=F{row-1}+E{row-1}"
            ws[f"G{row}"] = f"=2*F{row}"
            ws[f"H{row}"] = f"=$R$2-G{row}"
        total_row = last_data_row + 1
        ws[f"A{total_row}"] = "รวม"
        ws[f"B{total_row}"] = f"=SUM(B{first_data_row}:B{last_data_row})"
        ws[f"C{total_row}"] = f"=SUM(C{first_data_row}:C{last_data_row})"
        for cell in ws[total_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="F5F0E6")
        header_fill = PatternFill("solid", fgColor="0B2341")
        for cell in ws[9]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        widths = [9, 19, 19, 14, 14, 14, 14, 15, 17, 16, 22, 14, 12, 12, 28]
        for col, width in zip("ABCDEFGHIJKLMNO", widths):
            ws.column_dimensions[col].width = width
        versine_chart = LineChart()
        versine_chart.title = "เวอร์ไซน์เดิมและเวอร์ไซน์ใหม่"
        versine_chart.y_axis.title = "เวอร์ไซน์ (มม.)"
        versine_chart.x_axis.title = "หมุดที่"
        versine_chart.add_data(Reference(ws, min_col=2, max_col=3, min_row=9, max_row=last_data_row), titles_from_data=True)
        versine_chart.set_categories(Reference(ws, min_col=1, min_row=10, max_row=last_data_row))
        versine_chart.height = 8
        versine_chart.width = 16
        ws.add_chart(versine_chart, "Q12")
        throw_chart = LineChart()
        throw_chart.title = "ระยะดัด"
        throw_chart.y_axis.title = "ระยะดัด (มม.)"
        throw_chart.x_axis.title = "หมุดที่"
        throw_chart.add_data(Reference(ws, min_col=7, min_row=9, max_row=last_data_row), titles_from_data=True)
        throw_chart.set_categories(Reference(ws, min_col=1, min_row=10, max_row=last_data_row))
        throw_chart.height = 8
        throw_chart.width = 16
        ws.add_chart(throw_chart, "Q28")
    st.download_button(
        "ดาวน์โหลดตาราง String Lining พร้อมสูตรและกราฟ",
        output.getvalue(),
        file_name="string_lining_calculation.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


with st.sidebar:
    st.markdown("## SRT · CALC")
    st.caption("เครื่องมือคำนวณงานทางรถไฟ")
    selected = st.radio("หมวดคำนวณ", list(PAGES), label_visibility="collapsed")
    st.markdown("---")
    st.caption("อ้างอิงสูตรจากไฟล์ “รวมรายการคำนวณ ทางรถไฟ (วศธ.ทส.)”")
    st.caption("จัดทำระบบ: กองทางถาวร")

page = PAGES[selected]


if page == "overview":
    heading("เครื่องมือคำนวณงานทางรถไฟ", "เลือกหมวดจากแถบด้านซ้าย กรอกข้อมูล และส่งออกผลเป็น Excel ได้ทันที")
    c1, c2, c3 = st.columns(3)
    c1.metric("ชีตต้นฉบับ", "20", "รวมชีตซ่อน")
    c2.metric("เครื่องมือพร้อมใช้", "8", "สูตรหลักจาก Excel")
    c3.metric("ระบบหน่วย", "SI", "พร้อมหน่วยกำกับ")
    st.markdown("### หมวดที่พร้อมคำนวณ")
    ready = [
        ("Alignment", "ค่ายกโค้ง ความเร็วสูงสุด–ต่ำสุด โค้งต่อและโค้งกลม"),
        ("String Lining", "ตารางเวอร์ไซน์ ผลต่างสะสม ครึ่งระยะดัด ระยะดัด และกราฟสรุป"),
        ("Turnout", "ความเร็วผ่านประแจ เรขาคณิตประแจ และมุมตะเฆ้"),
        ("Track Structure", "แรงเค้นราง แรงถ่ายทอดสู่หมอน แรงกดหินโรยทางและพื้นทาง"),
        ("Drainage", "Rational Formula + Manning สำหรับรางรูปสี่เหลี่ยมคางหมู"),
        ("Gauge & Clearance", "ตรวจเรขาคณิตและระยะทางตรงจากการขยายขนาดทาง"),
    ]
    for title, body in ready:
        st.markdown(f'<div class="catalog-row"><b>{title}</b><br><span>{body}</span></div>', unsafe_allow_html=True)
    with st.expander("รายการชีตทั้งหมดจาก Excel ต้นฉบับ"):
        sheets = [
            "Track Design", "Drainage Design", "ข้อมูล INPUT", "หมอนคอนกรีต", "TQI ก่อน",
            "Normal distribution", "Turnout Design", "Turnout Design สรุป", "ความเร็วผ่านประแจ",
            "ความเร็วโค้ง", "คำนวณโค้งต่อ+กลม", "คำนวณโค้งด้วยพิกัด N,E", "มุมประแจ",
            "การขยายขนาดทาง จากราง", "การขยายจากแคร่รถจักร 3 เพลา", "String Lining ทส.1",
            "String Lining ทส.2", "40266ทส.", "ตาราง 3 แบบเลขที่ 4706-3", "Cant Calc",
        ]
        st.write(" · ".join(sheets))


elif page == "cant":
    heading("ค่ายกโค้งและความเร็ว", "คำนวณค่ายกสมดุล ส่วนขาด/ส่วนเกิน และช่วงความเร็วที่ยอมให้ผ่านโค้ง")
    left, right = st.columns([1, 1.45], gap="large")
    with left:
        radius = st.number_input("รัศมีโค้ง R (ม.)", 100.0, 10000.0, 1600.0, 10.0)
        speed = st.number_input("ความเร็วออกแบบ V (กม./ชม.)", 10.0, 250.0, 120.0, 5.0)
        auto = st.toggle("ใช้ค่ายกจริงที่ระบบแนะนำ", True)
        actual = None if auto else st.number_input("ค่ายกจริง Ca (มม.)", 0.0, 180.0, 55.0, 5.0)
        result = cant_calculation(radius, speed, actual)
        st.markdown('<div class="formula">Ce = 8.416V²/R · Ca = 5.611(V+5)²/R</div>', unsafe_allow_html=True)
    with right:
        st.markdown('<div class="section-label">ผลคำนวณ</div>', unsafe_allow_html=True)
        a, b, c = st.columns(3)
        a.metric("ค่ายกสมดุล Ce", f'{result["equilibrium_cant_mm"]:.0f} มม.')
        b.metric("ค่ายกจริง Ca", f'{result["actual_cant_mm"]:.0f} มม.')
        c.metric("ส่วนขาด Cd", f'{result["cant_deficiency_mm"]:.0f} มม.')
        a, b, c = st.columns(3)
        a.metric("ส่วนเกิน", f'{result["cant_excess_mm"]:.0f} มม.')
        b.metric("V สูงสุด", f'{result["vmax_kmh"]:.0f} กม./ชม.')
        c.metric("V ต่ำสุด", f'{result["vmin_kmh"]:.0f} กม./ชม.')
        result_status("Cant deficiency", result["deficiency_ok"], "เกณฑ์ใน Excel ≤ 50 มม.")
        result_status("Cant excess", result["excess_ok"], "เกณฑ์ใน Excel ≤ 60 มม.")
        speeds = list(range(20, 181, 5))
        chart = pd.DataFrame({"ความเร็ว (กม./ชม.)": speeds, "ค่ายกสมดุล (มม.)": [8.416 * v**2 / radius for v in speeds]})
        st.line_chart(chart, x="ความเร็ว (กม./ชม.)", y="ค่ายกสมดุล (มม.)", height=230)
        excel_download("cant_calculation", {"R (m)": radius, "V (km/h)": speed}, result)


elif page == "turnout_speed":
    heading("ความเร็วผ่านประแจทางเลี้ยว", "หาความเร็วจำกัดจากรัศมีและความเร่งด้านข้างที่เกิดจากส่วนขาดค่ายก")
    left, right = st.columns([1, 1.35], gap="large")
    with left:
        presets = {"1:8 · R 164.496 m": 164.496, "1:10 · R 185 m": 185.0, "1:12 · R 265 m": 265.0, "1:16 · R 648 m": 648.0, "กำหนดเอง": 300.0}
        preset = st.selectbox("แบบประแจ", presets)
        radius = st.number_input("รัศมีทางเลี้ยว (ม.)", 50.0, 3000.0, presets[preset], 1.0)
        hd = st.number_input("ส่วนขาดค่ายก hd (มม.)", 10.0, 100.0, 50.0, 5.0)
        gauge = st.number_input("ระยะศูนย์กลางราง G (มม.)", 900.0, 1600.0, 1067.0, 1.0)
        result = turnout_speed(radius, hd, gauge)
        st.markdown('<div class="formula">ac = g·hd/G · Vlim = 3.6√(ac·R)</div>', unsafe_allow_html=True)
    with right:
        a, b, c = st.columns(3)
        a.metric("ความเร่งด้านข้าง", f'{result["lateral_acceleration_mps2"]:.3f} m/s²')
        b.metric("V จำกัด", f'{result["limit_speed_kmh"]:.0f} กม./ชม.')
        c.metric("V ออกแบบ", f'{result["design_speed_kmh"]:.0f} กม./ชม.')
        st.info("ความเร็วออกแบบปัดลงทีละ 5 กม./ชม. ตามแนวทางใน Excel ต้นฉบับ")
        excel_download("turnout_speed", {"R (m)": radius, "hd (mm)": hd, "G (mm)": gauge}, result)


elif page == "drainage":
    heading("ออกแบบรางระบายน้ำข้างทาง", "คำนวณปริมาณน้ำด้วย Rational Formula และขนาดรางคางหมูด้วย Manning Formula")
    c1, c2, c3 = st.columns(3)
    with c1:
        length = st.number_input("ความยาวพื้นที่รับน้ำ (กม.)", .1, 100.0, 2.0, .1)
        width = st.number_input("ความกว้างพื้นที่รับน้ำ (ม.)", 1.0, 1000.0, 40.0, 1.0)
        intensity = st.number_input("ความเข้มฝน I (มม./ชม.)", 1.0, 500.0, 117.3, 1.0)
        runoff = st.number_input("สัมประสิทธิ์การไหลนอง C", 0.05, 1.0, .37, .01)
    with c2:
        n = st.number_input("Manning n", .005, .2, .03, .005, format="%.3f")
        slope = st.number_input("ความลาดชัน S", .0001, .2, .01, .001, format="%.4f")
        bottom = st.number_input("ท้องราง B (ม.)", .1, 20.0, .6, .1)
        z = st.number_input("ลาดข้าง z (H:V)", .0, 10.0, 1.0, .25)
        result = drainage_design(length, width, intensity, runoff, n, slope, bottom, z)
    with c3:
        st.metric("อัตราการไหล Q", f'{result["design_discharge_m3s"]:.3f} m³/s')
        st.metric("ความลึกน้ำ y", f'{result["flow_depth_m"]:.2f} ม.')
        st.metric("ความลึกรวม h", f'{result["total_depth_m"]:.2f} ม.')
        st.metric("ความกว้างปากราง T", f'{result["top_width_m"]:.2f} ม.')
        st.metric("ความเร็ว V", f'{result["velocity_mps"]:.2f} m/s')
        result_status("ไม่ตกตะกอน", result["sedimentation_ok"], "V ≥ 0.76 m/s")
        result_status("ไม่กัดเซาะ", result["erosion_ok"], "V ≤ 1.52 m/s")
        excel_download("drainage_design", {"L (km)": length, "width (m)": width, "I": intensity, "C": runoff, "n": n, "S": slope, "B": bottom, "z": z}, result)
    st.markdown('<div class="formula">Q = 0.278CIA · Manning Q = (1/n)AR^(2/3)S^(1/2)</div>', unsafe_allow_html=True)


elif page == "sleeper":
    heading("วิเคราะห์ราง–หมอน–หินโรยทาง", "ตรวจแรงเค้นในราง แรงถ่ายทอด และแรงกดบนชั้นหินโรยทางกับพื้นทาง")
    c1, c2, c3 = st.columns(3)
    with c1:
        rail = st.selectbox("ขนาดราง", list(RAIL_PROPERTIES), index=list(RAIL_PROPERTIES).index("BS 100 A"))
        ballast_c = st.number_input("สัมประสิทธิ์หินโรยทาง c (กก./ซม.³)", 1.0, 100.0, 10.0, 1.0)
        axle = st.number_input("น้ำหนักกดเพลา P (ตัน)", 1.0, 40.0, 20.0, 1.0)
        speed = st.number_input("ความเร็วสูงสุด (กม./ชม.)", 10.0, 250.0, 120.0, 5.0)
    with c2:
        sw = st.number_input("ความกว้างหมอน b (ซม.)", 10.0, 50.0, 25.0, 1.0)
        sl = st.number_input("ความยาวหมอน l (ซม.)", 100.0, 300.0, 200.0, 5.0)
        spacing = st.number_input("ระยะห่างหมอน a (ซม.)", 30.0, 100.0, 60.0, 5.0)
        thickness = st.number_input("ความหนาหินโรยทาง h (ซม.)", 10.0, 100.0, 25.0, 5.0)
        result = sleeper_track_analysis(rail, ballast_c, axle, sw, sl, spacing, thickness, speed)
    with c3:
        st.metric("แรงเค้นในราง σ", f'{result["rail_stress_ton_cm2"]:.3f} ตัน/ซม.²')
        st.metric("แรงสู่หมอน Prs", f'{result["rail_to_sleeper_ton"]:.2f} ตัน')
        st.metric("แรงกดหินโรยทาง Ps", f'{result["ballast_pressure_ton_m2"]:.2f} ตัน/ม.²')
        st.metric("แรงกดพื้นทาง Pb", f'{result["formation_pressure_ton_m2"]:.2f} ตัน/ม.²')
        result_status("แรงเค้นราง", result["rail_stress_ok"], "σa = 2.36 ตัน/ซม.²")
        result_status("หินโรยทาง", result["ballast_pressure_ok"], "Psa = 30 ตัน/ม.²")
        result_status("พื้นทาง", result["formation_pressure_ok"], "Pba = 15 ตัน/ม.²")
        excel_download("track_structure", {"rail": rail, "c": ballast_c, "axle": axle, "speed": speed, "b": sw, "l": sl, "spacing": spacing, "h": thickness}, result)


elif page == "turnout_geometry":
    heading("เรขาคณิตประแจ", "คำนวณมุมตะเฆ้ Theoretical lead และความยาวชุดประแจ")
    left, right = st.columns([1, 1.35], gap="large")
    with left:
        number = st.number_input("เบอร์ประแจ 1:N", 4.0, 40.0, 12.0, 1.0)
        straight = st.number_input("ทางตรงก่อนถึงจุด C (มม.)", 0.0, 5000.0, 1000.0, 50.0)
        radius = st.number_input("รัศมีโค้งประแจ (มม.)", 50000.0, 2000000.0, 264500.0, 500.0)
        tail = st.number_input("TNC ถึงท้ายตะเฆ้ e (มม.)", 0.0, 20000.0, 3252.0, 10.0)
        result = turnout_geometry(number, straight, radius, tail)
    with right:
        a, b = st.columns(2)
        a.metric("มุมตะเฆ้", f'{result["crossing_angle_deg"]:.4f}°')
        b.metric("PC–PI", f'{result["tangent_pc_pi_mm"]:,.0f} มม.')
        a.metric("Projection โค้ง", f'{result["curve_projection_mm"]:,.0f} มม.')
        b.metric("Lead เบื้องต้น", f'{result["preliminary_lead_mm"]:,.0f} มม.')
        st.info("Lead เบื้องต้นใช้ตรวจเรขาคณิตขั้นต้น ความยาวชุดประแจจริงต้องใช้มิติ switch panel, closure panel และ crossing panel จากแบบที่อนุมัติ")
        st.markdown('<div class="formula">θ = 2atan(0.5/N) · T = C/sinθ · Lead₀ = Rsinθ + Ccosθ</div>', unsafe_allow_html=True)
        excel_download("turnout_geometry", {"N": number, "C (mm)": straight, "R (mm)": radius, "e (mm)": tail}, result)


elif page == "transition":
    heading("โค้งต่อและโค้งกลม", "คำนวณความยาวโค้งต่อ ระยะ Shift และ Chainage ของ TS–SC–CS–ST")
    left, right = st.columns([1, 1.45], gap="large")
    with left:
        delta = st.number_input("มุมเบี่ยงเบน Δ (องศา)", .1, 179.0, 20.376667, .1, format="%.6f")
        radius = st.number_input("รัศมีโค้ง R (ม.)", 50.0, 10000.0, 600.0, 10.0)
        speed = st.number_input("ความเร็วออกแบบ V (กม./ชม.)", 10.0, 250.0, 80.0, 5.0)
        pi_ch = st.number_input("Chainage PI (ม.)", 0.0, 5000000.0, 700514.123, 10.0, format="%.3f")
        result = transition_curve(delta, radius, speed, pi_ch)
    with right:
        a, b, c = st.columns(3)
        a.metric("ค่ายกจริง", f'{result["actual_cant_mm"]:.0f} มม.')
        b.metric("โค้งต่อ Ls", f'{result["transition_length_m"]:.2f} ม.')
        c.metric("Shift", f'{result["shift_m"]:.3f} ม.')
        chainages = pd.DataFrame({"จุด": ["TS", "SC", "CS", "ST"], "Chainage (m)": [result["ts_chainage_m"], result["sc_chainage_m"], result["cs_chainage_m"], result["st_chainage_m"]]})
        st.dataframe(chainages.style.format({"Chainage (m)": "{:,.3f}"}), hide_index=True, use_container_width=True)
        st.markdown('<div class="formula">Ls = 0.01VCa · Shift = Ls²/(24R)</div>', unsafe_allow_html=True)
        excel_download("transition_curve", {"delta": delta, "R": radius, "V": speed, "PI": pi_ch}, result)


elif page == "string_lining":
    heading(
        "การคำนวณโค้งเส้นเชือก (String Lining)",
        "กรอกเวอร์ไซน์เดิมและเวอร์ไซน์แก้ไข ระบบคำนวณผลต่างสะสม ครึ่งระยะดัด ระยะดัด และกราฟรูปโค้ง",
    )

    with st.expander("ข้อมูลหัวแบบฟอร์ม", expanded=True):
        f1, f2, f3, f4 = st.columns(4)
        curve_no = f1.text_input("โค้งเลขที่", "40267")
        km_start = f2.text_input("กม. เริ่มต้น", "754+185")
        km_end = f3.text_input("กม. ปลายโค้ง", "754+320")
        direction = f4.selectbox("ทิศทางโค้ง", ["โค้งขวา", "โค้งซ้าย"])
        f1, f2, f3 = st.columns(3)
        inspector_section = f1.text_input("ตอนนายตรวจทาง", "ทุ่งสง")
        division = f2.text_input("แขวงบำรุงทาง", "ทุ่งสง")
        office = f3.text_input("กองบำรุงทางเขต", "ทุ่งสง")
        f1, f2, f3, f4 = st.columns(4)
        speed_form = f1.number_input("ความเร็วขบวนรถ (กม./ชม.)", 10.0, 250.0, 80.0, 5.0)
        radius = f2.number_input("รัศมีเป้าหมาย R (ม.)", 50.0, 10000.0, 800.0, 10.0)
        cant_form = f3.number_input("ค่ายกโค้ง (มม.)", 0.0, 180.0, 50.0, 5.0)
        curve_length_form = f4.number_input("ความยาวโค้งรวม (ม.)", 1.0, 10000.0, 135.0, 5.0)
        f1, f2, f3, f4 = st.columns(4)
        surveyor = f1.text_input("ผู้สำรวจ", "นตท.")
        survey_date = f2.date_input("วันที่สำรวจ", value=date.today())
        calculator_name = f3.text_input("ผู้คำนวณ", "วศธ.ทส.")
        calculation_date = f4.date_input("วันที่คำนวณ", value=date.today())
        f1, f2 = st.columns(2)
        centre_setter = f1.text_input("ผู้ให้ ศก.หมุด", "")
        lining_operator = f2.text_input("ผู้ดัดราง", "")

    metadata = {
        "curve_no": curve_no,
        "km_start": km_start,
        "km_end": km_end,
        "direction": direction,
        "inspector_section": inspector_section,
        "division": division,
        "office": office,
        "speed": speed_form,
        "radius": radius,
        "cant": cant_form,
        "curve_length": curve_length_form,
        "surveyor": surveyor,
        "survey_date": survey_date.isoformat(),
        "calculator": calculator_name,
        "calculation_date": calculation_date.isoformat(),
        "centre_setter": centre_setter,
        "lining_operator": lining_operator,
    }

    st.markdown("### ผังองค์ประกอบโค้ง")
    string_lining_curve_diagram(direction)
    st.caption("TS: Tangent to Spiral · SC: Spiral to Circular · CS: Circular to Spiral · ST: Spiral to Tangent")

    sample_original = [6, 1, 5, 10, 7, 0, 2, 18, 14, 17, 17, 15, 11, 16, 16, 16, 16, 15, 13, 11, 6, 9, 8, 7, 4, 3, 0, 0, 0, 0, 5]
    sample_revised = [0, 1, 3, 6, 8, 9, 10, 12, 14, 15, 16, 16, 16, 16, 15, 14, 15, 15, 13, 13, 11, 8, 7, 5, 4, 3, 2, 1, 0, 0, 0]
    sample_notes = [""] * len(sample_original)
    for pin, note in {1: "TS", 10: "SC", 18: "CS", 28: "ST"}.items():
        sample_notes[pin - 1] = note

    extra_defaults = {
        "ค่าขยายทาง (มม.)": 0.0,
        "ค่ายกโค้ง (มม.)": 0.0,
        "เวอร์ไซน์มาตรฐาน (มม.)": 0.0,
        "สทล. จุดที่": "",
        "ศก.ทาง": "",
        "รางนอก": "",
    }

    def complete_string_lining_columns(frame: pd.DataFrame) -> pd.DataFrame:
        completed = frame.copy()
        for column, default in extra_defaults.items():
            if column not in completed:
                completed[column] = default
        columns = [
            "หมุดที่", "เวอร์ไซน์เดิม (มม.)", "เวอร์ไซน์ใหม่ (มม.)",
            "ค่าขยายทาง (มม.)", "ค่ายกโค้ง (มม.)", "เวอร์ไซน์มาตรฐาน (มม.)",
            "สทล. จุดที่", "ศก.ทาง", "รางนอก", "หมายเหตุ",
        ]
        return completed[columns]

    if "string_lining_input" not in st.session_state:
        st.session_state.string_lining_input = complete_string_lining_columns(pd.DataFrame(
            {
                "หมุดที่": range(1, len(sample_original) + 1),
                "เวอร์ไซน์เดิม (มม.)": sample_original,
                "เวอร์ไซน์ใหม่ (มม.)": sample_revised,
                "หมายเหตุ": sample_notes,
            }
        ))
    else:
        st.session_state.string_lining_input = complete_string_lining_columns(st.session_state.string_lining_input)

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        chord_length = st.number_input("ความยาวคอร์ด C (ม.)", 1.0, 40.0, 10.0, 1.0)
    with c2:
        pin_spacing = st.number_input("ระยะหมุด/ครึ่งคอร์ด (ม.)", .5, 20.0, 5.0, .5)
    with c3:
        baseline = st.number_input("ระยะฐานราง-ศก. (มม.)", 0.0, 5000.0, 500.0, 10.0)
    with c4:
        target_throw = st.select_slider("เป้าหมายระยะดัดสูงสุด (มม.)", options=[60, 70, 80, 90, 100, 110, 120], value=120)

    theoretical_max = chord_length**2 * 1000 / (8 * radius)
    st.caption(f"เวอร์ไซน์ทฤษฎีในโค้งกลม V = C²/(8R) = {theoretical_max:.2f} มม. · ตอกหมุดตามแนวรางทุกครึ่งคอร์ด = {pin_spacing:g} ม.")
    with st.expander("หลักเกณฑ์จากคู่มือการดัดแนวโค้ง"):
        st.markdown(
            """
- ใช้เชือกยาวเท่ากันตลอดแนว และวัดเวอร์ไซน์ที่กึ่งกลางคอร์ด
- หมุดแรกและหมุดสุดท้ายควรอยู่บนทางตรง โดยเวอร์ไซน์แก้ไขเป็นศูนย์
- เวอร์ไซน์ในโค้งต่อควรเพิ่มหรือลดแบบ Arithmetic Progression และในโค้งกลมควรสม่ำเสมอ
- ผลรวมเวอร์ไซน์เดิมต้องเท่ากับผลรวมเวอร์ไซน์ใหม่ รวมต่างและครึ่งระยะดัดที่ปลายต้องกลับเป็นศูนย์
- ค่าระยะดัดสูงสุดควรไม่เกิน 20 ซม. หากทำได้ และต้องตรวจสิ่งกีดขวาง/ระยะปลอดภัยก่อนดัดจริง
"""
        )

    with st.expander("สร้างเวอร์ไซน์ใหม่เบื้องต้นแบบ Arithmetic Progression"):
        p1, p2, p3, p4, p5 = st.columns(5)
        before = p1.number_input("ทางตรงก่อน TS (หมุด)", 0, 50, 2)
        entry = p2.number_input("โค้งต่อด้านต้น (ช่วง)", 1, 100, 9)
        circular = p3.number_input("โค้งกลม (หมุด)", 0, 200, 9)
        exit_count = p4.number_input("โค้งต่อด้านปลาย (ช่วง)", 1, 100, 9)
        after = p5.number_input("ทางตรงหลัง ST (หมุด)", 0, 50, 2)
        st.caption("ค่าที่สร้างเป็นจุดเริ่มต้นสำหรับปรับแก้ ต้องตรวจผลรวมและการปิดค่าท้ายตารางอีกครั้ง")
        b1, b2 = st.columns(2)
        if b1.button("สร้างโปรไฟล์เวอร์ไซน์ใหม่", use_container_width=True):
            generated = theoretical_versine_profile(before, entry, circular, exit_count, after, theoretical_max)
            current = st.session_state.string_lining_input
            old = current["เวอร์ไซน์เดิม (มม.)"].tolist() if len(current) == len(generated) else [0.0] * len(generated)
            st.session_state.string_lining_input = complete_string_lining_columns(pd.DataFrame(
                {"หมุดที่": range(1, len(generated) + 1), "เวอร์ไซน์เดิม (มม.)": old, "เวอร์ไซน์ใหม่ (มม.)": generated, "หมายเหตุ": [""] * len(generated)}
            ))
            st.session_state.pop("string_lining_editor", None)
            st.rerun()
        if b2.button("เรียกตัวอย่างจาก Excel ทส.1", use_container_width=True):
            st.session_state.string_lining_input = complete_string_lining_columns(pd.DataFrame(
                {"หมุดที่": range(1, len(sample_original) + 1), "เวอร์ไซน์เดิม (มม.)": sample_original, "เวอร์ไซน์ใหม่ (มม.)": sample_revised, "หมายเหตุ": sample_notes}
            ))
            st.session_state.pop("string_lining_editor", None)
            st.rerun()

    st.markdown("### 1. ตารางกรอกข้อมูล")
    st.caption("แก้ไขเฉพาะเวอร์ไซน์เดิม เวอร์ไซน์ใหม่ และหมายเหตุ สามารถเพิ่มหรือลบแถวได้")
    edited = st.data_editor(
        st.session_state.string_lining_input,
        num_rows="dynamic",
        hide_index=True,
        use_container_width=True,
        column_config={
            "หมุดที่": st.column_config.NumberColumn("หมุดที่", min_value=1, step=1, format="%d"),
            "เวอร์ไซน์เดิม (มม.)": st.column_config.NumberColumn("เวอร์ไซน์เดิม (มม.)", step=1.0, format="%.1f"),
            "เวอร์ไซน์ใหม่ (มม.)": st.column_config.NumberColumn("เวอร์ไซน์ใหม่ (มม.)", step=1.0, format="%.1f"),
            "ค่าขยายทาง (มม.)": st.column_config.NumberColumn("ค่าขยายทาง (มม.)", step=1.0, format="%.1f"),
            "ค่ายกโค้ง (มม.)": st.column_config.NumberColumn("ค่ายกโค้ง (มม.)", step=1.0, format="%.1f"),
            "เวอร์ไซน์มาตรฐาน (มม.)": st.column_config.NumberColumn("เวอร์ไซน์มาตรฐาน (มม.)", step=1.0, format="%.1f"),
            "สทล. จุดที่": st.column_config.TextColumn("สทล. จุดที่"),
            "ศก.ทาง": st.column_config.TextColumn("ศก.ทาง"),
            "รางนอก": st.column_config.TextColumn("รางนอก"),
            "หมายเหตุ": st.column_config.TextColumn("หมายเหตุ"),
        },
        key="string_lining_editor",
    )
    working = edited.copy()
    working["เวอร์ไซน์เดิม (มม.)"] = pd.to_numeric(working["เวอร์ไซน์เดิม (มม.)"], errors="coerce")
    working["เวอร์ไซน์ใหม่ (มม.)"] = pd.to_numeric(working["เวอร์ไซน์ใหม่ (มม.)"], errors="coerce")
    working = working.dropna(subset=["เวอร์ไซน์เดิม (มม.)", "เวอร์ไซน์ใหม่ (มม.)"]).reset_index(drop=True)
    working["หมุดที่"] = range(1, len(working) + 1)
    working["หมายเหตุ"] = working["หมายเหตุ"].fillna("")
    for column in ["ค่าขยายทาง (มม.)", "ค่ายกโค้ง (มม.)", "เวอร์ไซน์มาตรฐาน (มม.)"]:
        working[column] = pd.to_numeric(working[column], errors="coerce").fillna(0.0)
    for column in ["สทล. จุดที่", "ศก.ทาง", "รางนอก"]:
        working[column] = working[column].fillna("")
    st.session_state.string_lining_input = working
    if working.empty:
        st.warning("กรุณาเพิ่มข้อมูลเวอร์ไซน์อย่างน้อย 1 หมุด")
        st.stop()

    action_left, action_right = st.columns([2, 1])
    if action_left.button("ปรับปรุง Vใหม่ ให้ผ่านเงื่อนไข", type="primary", use_container_width=True):
        try:
            improved = improve_string_lining_versines(
                working["เวอร์ไซน์เดิม (มม.)"].tolist(),
                working["เวอร์ไซน์ใหม่ (มม.)"].tolist(),
                target_throw,
            )
            st.session_state.string_lining_previous_revised = working["เวอร์ไซน์ใหม่ (มม.)"].tolist()
            working["เวอร์ไซน์ใหม่ (มม.)"] = improved["revised_versines"]
            st.session_state.string_lining_input = working
            target_message = (
                "อยู่ในเป้าหมาย" if improved["maximum_throw_mm"] <= target_throw
                else "ยังเกินเป้าหมาย กรุณาตรวจข้อจำกัดหน้างานและปรับด้วยวิศวกร"
            )
            st.session_state.string_lining_optimization_message = (
                f'ปรับ Vใหม่แล้ว: ปิดผลรวมและครึ่งระยะดัด ระยะดัดสูงสุด {improved["maximum_throw_mm"]:.1f} มม. - {target_message}'
            )
            st.session_state.pop("string_lining_editor", None)
            st.rerun()
        except ValueError as error:
            st.error(str(error))
    if action_right.button("ย้อนกลับ Vใหม่", use_container_width=True, disabled="string_lining_previous_revised" not in st.session_state):
        previous = st.session_state.string_lining_previous_revised
        if len(previous) == len(working):
            working["เวอร์ไซน์ใหม่ (มม.)"] = previous
            st.session_state.string_lining_input = working
            st.session_state.pop("string_lining_previous_revised", None)
            st.session_state.pop("string_lining_editor", None)
            st.rerun()
    if "string_lining_optimization_message" in st.session_state:
        st.success(st.session_state.pop("string_lining_optimization_message"))

    result = string_lining_calculation(
        working["เวอร์ไซน์เดิม (มม.)"].tolist(),
        working["เวอร์ไซน์ใหม่ (มม.)"].tolist(),
        baseline,
    )
    result_df = pd.DataFrame(
        {
            "หมุดที่": [row["station"] for row in result["rows"]],
            "เวอร์ไซน์เดิม (มม.)": [row["original_versine_mm"] for row in result["rows"]],
            "เวอร์ไซน์ใหม่ (มม.)": [row["revised_versine_mm"] for row in result["rows"]],
            "ผลต่าง (มม.)": [row["difference_mm"] for row in result["rows"]],
            "รวมต่าง (มม.)": [row["cumulative_difference_mm"] for row in result["rows"]],
            "ครึ่งดัด (มม.)": [row["half_throw_mm"] for row in result["rows"]],
            "ระยะดัด (มม.)": [row["throw_mm"] for row in result["rows"]],
            "ราง-ศก. (มม.)": [row["rail_to_centre_mm"] for row in result["rows"]],
            "ค่าขยายทาง (มม.)": working["ค่าขยายทาง (มม.)"],
            "ค่ายกโค้ง (มม.)": working["ค่ายกโค้ง (มม.)"],
            "เวอร์ไซน์มาตรฐาน (มม.)": working["เวอร์ไซน์มาตรฐาน (มม.)"],
            "สทล. จุดที่": working["สทล. จุดที่"],
            "ศก.ทาง": working["ศก.ทาง"],
            "รางนอก": working["รางนอก"],
            "หมายเหตุ": working["หมายเหตุ"],
        }
    )

    st.markdown("### 2. ผลการคำนวณ")
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Σ เวอร์ไซน์เดิม", f'{result["sum_original_mm"]:.0f} มม.')
    k2.metric("Σ เวอร์ไซน์ใหม่", f'{result["sum_revised_mm"]:.0f} มม.', f'{result["sum_difference_mm"]:+.0f} มม.')
    k3.metric("ระยะดัดสูงสุด", f'{result["max_abs_throw_mm"]:.0f} มม.')
    k4.metric("ปิดครึ่งดัดปลาย", f'{result["final_half_throw_mm"]:+.0f} มม.')
    st.dataframe(
        result_df.style.format({col: "{:.1f}" for col in result_df.columns if "(มม.)" in col}),
        hide_index=True,
        use_container_width=True,
        height=440,
    )

    st.markdown("### 3. กราฟสรุป")
    graph_left, graph_right = st.columns(2, gap="large")
    with graph_left:
        st.caption("รูปโค้งจากเวอร์ไซน์เดิมและเวอร์ไซน์ใหม่")
        st.line_chart(
            result_df[["หมุดที่", "เวอร์ไซน์เดิม (มม.)", "เวอร์ไซน์ใหม่ (มม.)"]],
            x="หมุดที่",
            y=["เวอร์ไซน์เดิม (มม.)", "เวอร์ไซน์ใหม่ (มม.)"],
            height=320,
        )
    with graph_right:
        st.caption("ระยะดัดที่แต่ละหมุด: ค่าบวก/ลบแสดงทิศทางการดัดคนละด้าน")
        st.line_chart(result_df[["หมุดที่", "ระยะดัด (มม.)"]], x="หมุดที่", y="ระยะดัด (มม.)", height=320)

    st.markdown("### 4. ตรวจสอบเงื่อนไข")
    q1, q2 = st.columns(2)
    with q1:
        result_status("ผลรวมเวอร์ไซน์", result["sum_balanced"], "ΣVเดิม ต้องเท่ากับ ΣVใหม่")
        result_status("รวมต่างหมุดสุดท้าย", result["cumulative_closed"], "ต้องเท่ากับศูนย์")
        result_status("ครึ่งระยะดัดปลายโค้ง", result["throw_closed"], "ต้องเริ่มและจบด้วยศูนย์")
    with q2:
        result_status("เวอร์ไซน์ปลายทั้งสอง", result["end_versines_zero"], "ทางตรงควรเป็นศูนย์")
        result_status("ระยะดัดเป้าหมาย", result["max_abs_throw_mm"] <= target_throw, f"กำหนดไม่เกิน {target_throw} มม. (เลือกได้ในช่วง 60-120 มม.)")
        result_status("ระยะดัดสูงสุดตามคู่มือ", result["throw_within_document_limit"], "ควรไม่เกิน 20 ซม. หากทำได้")

    st.markdown('<div class="formula">Dᵢ = Vเดิม−Vใหม่ · Eᵢ = Eᵢ₋₁+Dᵢ · Fᵢ = Fᵢ₋₁+Eᵢ₋₁ · ระยะดัด = 2Fᵢ</div>', unsafe_allow_html=True)
    string_lining_excel_download(working, result_df, baseline, metadata)


elif page == "widening":
    heading("การขยายขนาดทางจากราง", "ตรวจระยะทางตรงระหว่างชุดประแจและเรขาคณิตจากมุม 1:N")
    left, right = st.columns([1, 1.35], gap="large")
    with left:
        number = st.number_input("มุมประแจ 1:N", 4.0, 40.0, 12.0, 1.0)
        track_distance = st.number_input("ระยะระหว่างแนวทาง (ม.)", 2.0, 20.0, 4.4, .1)
        speed = st.number_input("ความเร็ว (กม./ชม.)", 10.0, 200.0, 50.0, 5.0)
        track_width = st.number_input("Track width W (มม.)", 900.0, 2000.0, 1500.0, 10.0)
        hd = st.number_input("Cant deficiency ที่ยอมให้ (มม.)", 10.0, 200.0, 90.0, 5.0)
        result = track_widening(number, track_distance, speed, track_width, hd)
    with right:
        a, b, c = st.columns(3)
        a.metric("รัศมีใช้คำนวณ", f'{result["radius_m"]:.0f} ม.')
        b.metric("ระยะทางตรงที่มี", f'{result["available_straight_m"]:.2f} ม.')
        c.metric("ระยะที่ต้องการ", f'{result["required_straight_m"]:.2f} ม.')
        result_status("เรขาคณิตทางตรง", result["geometry_ok"], "ต้องไม่น้อยกว่า 0.2V")
        st.dataframe(pd.DataFrame({"รายการ": ["มุมประแจ", "Throw", "Tangent", "แนวทแยง"], "ค่า": [f'{result["angle_deg"]:.3f}°', f'{result["throw_m"]:.3f} m', f'{result["tangent_m"]:.3f} m', f'{result["diagonal_m"]:.3f} m']}), hide_index=True, use_container_width=True)
        excel_download("track_widening", {"N": number, "distance": track_distance, "V": speed, "W": track_width, "hd": hd}, result)

st.markdown("---")
st.caption("หมายเหตุ: ผลคำนวณเป็นเครื่องมือช่วยงานวิศวกรรม ต้องตรวจสอบข้อมูลสำรวจ แบบ มาตรฐานฉบับปัจจุบัน และให้วิศวกรผู้รับผิดชอบทบทวนก่อนนำไปใช้จริง")
